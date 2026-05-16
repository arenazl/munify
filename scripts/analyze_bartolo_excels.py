"""Analiza los 3 Excels de Bartolo y devuelve un reporte agregado."""
import openpyxl, re, unicodedata
from collections import defaultdict, Counter

FILES = [
    ('docs/bartolo/Gastos 2024.xlsx', 2024),
    ('docs/bartolo/Gastos mensuales 2025 - 01-09.xlsx', 2025),
    ('docs/bartolo/Gatos municipalidad oct-nov-dic.xlsx', 2025),
]


def norm_name(s):
    if not s:
        return ''
    s = str(s).strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = re.sub(r'\s*\(.*?\)\s*', ' ', s)
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    parts = sorted(s.split())
    return ' '.join(parts)


def norm_modo(s):
    if not s:
        return ''
    s = str(s).strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = re.sub(r'\s+', ' ', s).strip()
    if 'echeq' in s:
        return 'echeq'
    if 'cheque' in s:
        return 'cheque'
    if 'efect' in s or 'eftv' in s:
        return 'efectivo'
    if 'trans' in s:
        return 'transferencia'
    if s == 'mp' or s.startswith('mp ') or ' mp' in (' ' + s):
        return 'mercadopago'
    if 'mercado' in s:
        return 'mercadopago'
    if 'naran' in s:
        return 'naranja'
    if 'debito' in s:
        return 'debito'
    return s


def parse_amount(v):
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace('.', '').replace(',', '.').replace(' ', '')
    s = re.sub(r'[^\d.\-]', '', s)
    try:
        return float(s) if s else 0.0
    except Exception:
        return 0.0


sheets_per_file = defaultdict(list)
contacts_by_cat = defaultdict(set)
contact_display = {}
proyectos = set()
cajas = Counter()
cajas_monto = defaultdict(float)
gastos_count = 0
gastos_breakdown = []
gastos_por_anio = Counter()
empleados_blanco_negro = Counter()
modalidad_pago = Counter()

CONCEPTS = {
    'Pago de sueldos y jornales': ['sueldo', 'jornal'],
    'Pago de aguinaldo / SAC': ['aguinaldo', 'sac'],
    'Pago de honorarios profesionales': ['abogado', 'contador', 'ingeniero', 'arquitecto', 'escribano', 'honorario', 'psicolog', 'dr.', 'dra.', 'lic.'],
    'Pago de viáticos y movilidad': ['viatico', 'movilidad', 'traslado'],
    'Pago de servicios públicos (luz, agua, gas)': ['epec', ' luz', 'aguas cordobes', 'energia electrica', 'gas natural'],
    'Pago de Internet y telefonía': ['internet', 'telefon', 'claro', 'movistar', 'tigo', 'wifi', 'fibra'],
    'Pago de alquileres': ['alquiler', 'renta '],
    'Pago de seguros': ['seguro', 'aseguradora', 'sancor seguro', 'provincia seguro'],
    'Pago de impuestos y tasas': ['afip', 'rentas', 'impuesto', 'iibb', 'arba', 'dgip', 'sellos'],
    'Pago de gastos bancarios': ['gasto banc', 'comision banc', 'bancari'],
    'Pago de préstamos / devoluciones': ['prestamo', 'devolucion', 'cuota'],
    'Pagos varios': [],
    'Compra de combustible': ['ypf', 'combustible', 'nafta', 'gasoil', 'axion', 'shell', 'petrobras', 'puma', 'dapsa'],
    'Compra de materiales de obra': ['cemento', 'hierro', 'arena', 'ladrillo', 'corralon', 'santa rita', 'ferreteria', 'aceros', 'acerco'],
    'Compra de materiales de oficina': ['libreria', 'resma', 'tinta', 'oficina'],
    'Compra de insumos de limpieza': ['limpieza', 'detergente', 'lavandina'],
    'Compra de herramientas y equipamiento': ['herramienta', 'maquinaria', 'equipamiento', 'xgov'],
    'Compras varias': [],
    'Reparación de vehículos': ['lubricentro', 'mecanic', 'repuesto', 'neumatico', 'taller', 'cubierta', 'aceite '],
    'Reparación de edificios e instalaciones': ['albañil', 'pintura', 'reparacion edificio'],
    'Reparación de equipos': ['reparacion equipo', 'service '],
    'Reparaciones varias': ['reparacion', 'arreglo'],
    'Contratación de fletes y transporte': ['flete', 'transporte', 'tranporte', 'envio', 'camion', 'remis'],
    'Contratación de eventos y actividades culturales': ['evento', 'musico', 'banda', 'show', 'feria', 'fiesta', 'carpa', 'sonido', 'dj '],
    'Contratación de servicios profesionales': ['consultor', 'asesoria'],
    'Contrataciones varias': [],
    'Aporte a salud / prestaciones médicas': ['farmacia', 'enfermera', 'hospital', 'clinica', 'remedio', 'medicamento', 'salud'],
    'Aporte a subsidios y ayudas sociales': ['ayuda social', 'subsidio', 'asistencia social'],
    'Aportes varios': [],
    'Obra pública / construcción': ['vivienda semilla', 'balneario', 'pavimento', 'cordon cuneta', 'obra publica'],
    'Obras varias': ['obra'],
    'Otros gastos': [],
}

ORDER = ['Pago de aguinaldo / SAC', 'Pago de honorarios profesionales',
         'Compra de combustible', 'Pago de servicios públicos (luz, agua, gas)',
         'Pago de Internet y telefonía', 'Pago de seguros',
         'Pago de impuestos y tasas', 'Pago de gastos bancarios',
         'Reparación de vehículos', 'Contratación de fletes y transporte',
         'Contratación de eventos y actividades culturales',
         'Aporte a salud / prestaciones médicas',
         'Aporte a subsidios y ayudas sociales',
         'Compra de materiales de obra', 'Compra de materiales de oficina',
         'Compra de insumos de limpieza',
         'Compra de herramientas y equipamiento',
         'Obra pública / construcción', 'Obras varias',
         'Pago de alquileres', 'Pago de viáticos y movilidad',
         'Pago de sueldos y jornales',
         'Reparación de edificios e instalaciones',
         'Reparación de equipos', 'Reparaciones varias',
         'Contratación de servicios profesionales']


def map_concept(text):
    if not text:
        return 'Otros gastos'
    t = str(text).lower()
    t = ''.join(c for c in unicodedata.normalize('NFKD', t) if not unicodedata.combining(c))
    for c in ORDER:
        for kw in CONCEPTS[c]:
            if kw in t:
                return c
    return 'Otros gastos'


PROYECTO_KW = ['vivienda semilla', 'balneario', 'plaza', 'camping',
               'polideport', 'escuela', 'hospital', 'cementerio', 'salon',
               'centro cultural', 'cordon cuneta', 'pavimento']


def extract_proyecto(text):
    if not text:
        return None
    t = str(text).lower()
    t = ''.join(c for c in unicodedata.normalize('NFKD', t) if not unicodedata.combining(c))
    m = re.search(r'\(([^)]+)\)', t)
    if m:
        inner = m.group(1).strip()
        for kw in PROYECTO_KW:
            if kw in inner:
                return inner
    for kw in PROYECTO_KW:
        if kw in t:
            idx = t.find(kw)
            return t[max(0, idx - 3):idx + len(kw) + 12].strip()
    return None


concept_counts = Counter()

for fpath, default_year in FILES:
    wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
    for sh in wb.sheetnames:
        sheets_per_file[fpath].append(sh)
        ws = wb[sh]
        sh_low = sh.strip().lower()
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        gastos_sheet_count = 0
        gastos_sheet_total = 0.0

        if sh_low in ('octubre', 'noviembre', 'diciembre'):
            header = [str(c).lower().strip() if c else '' for c in rows[0]]

            def idx_of(opts):
                for i, h in enumerate(header):
                    if h in opts:
                        return i
                return -1
            i_monto = idx_of(['monto', 'gastos'])
            i_emp = idx_of(['empresa', 'proveedor'])
            i_modo = idx_of(['modo', 'forma', 'medio'])
            for r in rows[1:]:
                if not r:
                    continue
                monto = parse_amount(r[i_monto]) if 0 <= i_monto < len(r) else 0
                emp = r[i_emp] if 0 <= i_emp < len(r) else ''
                modo = r[i_modo] if 0 <= i_modo < len(r) else ''
                if monto > 0 and emp:
                    gastos_count += 1
                    gastos_sheet_count += 1
                    gastos_sheet_total += monto
                    gastos_por_anio[default_year] += 1
                    if modo:
                        m = norm_modo(modo)
                        cajas[m] += 1
                        cajas_monto[m] += monto
                    nm = norm_name(emp)
                    if nm:
                        contacts_by_cat['Proveedores / contratistas'].add(nm)
                        contact_display.setdefault(nm, str(emp).strip())
                    p = extract_proyecto(emp)
                    if p:
                        proyectos.add(p)
                    concept_counts[map_concept(emp)] += 1

        elif sh_low == 'gastos':
            hdr_idx = None
            for i, r in enumerate(rows[:6]):
                rs = [str(c).lower().strip() if c else '' for c in r]
                if 'empresa' in rs and 'modo' in rs:
                    hdr_idx = i
                    break
            if hdr_idx is None:
                continue
            header = [str(c).lower().strip() if c else '' for c in rows[hdr_idx]]
            empresa_idxs = [i for i, h in enumerate(header) if h == 'empresa']
            modo_idxs = [i for i, h in enumerate(header) if h == 'modo']
            blocks = []
            for ei in empresa_idxs:
                mi = ei - 1
                mo = next((m for m in modo_idxs if m > ei), ei + 1)
                blocks.append((mi, ei, mo))
            for r in rows[hdr_idx + 1:]:
                if not r:
                    continue
                for (mi, ei, mo) in blocks:
                    monto = parse_amount(r[mi]) if mi < len(r) else 0
                    emp = r[ei] if ei < len(r) else ''
                    modo = r[mo] if mo < len(r) else ''
                    if monto > 0 and emp and str(emp).strip():
                        gastos_count += 1
                        gastos_sheet_count += 1
                        gastos_sheet_total += monto
                        gastos_por_anio[default_year] += 1
                        if modo:
                            m_ = norm_modo(modo)
                            cajas[m_] += 1
                            cajas_monto[m_] += monto
                        nm = norm_name(emp)
                        if nm:
                            contacts_by_cat['Proveedores / contratistas'].add(nm)
                            contact_display.setdefault(nm, str(emp).strip())
                        p = extract_proyecto(emp)
                        if p:
                            proyectos.add(p)
                        concept_counts[map_concept(emp)] += 1

        elif sh_low == 'empleados':
            for r in rows[1:]:
                if not r or len(r) < 3:
                    continue
                name = r[2]
                if not name:
                    continue
                nm = norm_name(name)
                if not nm or nm in ('empleado', 'trabajo'):
                    continue
                contacts_by_cat['Empleados de planta'].add(nm)
                contact_display.setdefault(nm, str(name).strip())
                col_e = str(r[4]).lower().strip() if len(r) > 4 and r[4] else ''
                if 'negro' in col_e and 'blanco' in col_e:
                    empleados_blanco_negro['blanco y negro'] += 1
                elif 'negro' in col_e:
                    empleados_blanco_negro['negro'] += 1
                elif 'blanco' in col_e:
                    empleados_blanco_negro['blanco'] += 1
                row_text = ' '.join(str(c).lower() for c in r if c)
                if 'jubilad' in row_text:
                    modalidad_pago['Jubilado'] += 1
                elif 'trafic' in row_text:
                    modalidad_pago['Trafic'] += 1
                elif 'transfer' in row_text or 'transf' in row_text:
                    modalidad_pago['Transferencia'] += 1
                elif 'recibo' in row_text:
                    modalidad_pago['Recibo'] += 1

        elif sh_low == 'concejales':
            for r in rows[2:]:
                if not r:
                    continue
                name = r[0]
                if not name or str(name).strip().lower() in ('concejales', ''):
                    continue
                nm = norm_name(name)
                if not nm:
                    continue
                contacts_by_cat['Concejales'].add(nm)
                contact_display.setdefault(nm, str(name).strip())

        elif 'turismo' in sh_low and 'cultura' in sh_low:
            for r in rows[1:]:
                if not r:
                    continue
                name = r[0]
                if not name:
                    continue
                nm = norm_name(name)
                if not nm:
                    continue
                contacts_by_cat['Personal de Turismo y cultura'].add(nm)
                contact_display.setdefault(nm, str(name).strip())

        elif 'prof' in sh_low and ('pub' in sh_low or 'publicidad' in sh_low):
            for r in rows[1:]:
                if not r:
                    continue
                name = r[0]
                if not name:
                    continue
                nm = norm_name(name)
                if not nm:
                    continue
                contacts_by_cat['Profesionales y publicidad'].add(nm)
                contact_display.setdefault(nm, str(name).strip())

        elif 'ayuda social' in sh_low:
            for r in rows[1:]:
                if not r:
                    continue
                name = r[0]
                if not name:
                    continue
                nm = norm_name(name)
                if not nm:
                    continue
                contacts_by_cat['Ayuda social'].add(nm)
                contact_display.setdefault(nm, str(name).strip())

        elif 'gerontol' in sh_low:
            for r in rows[2:]:
                if not r:
                    continue
                name = r[0]
                if not name:
                    continue
                nm = norm_name(name)
                if not nm:
                    continue
                contacts_by_cat['Apoyo gerontológico'].add(nm)
                contact_display.setdefault(nm, str(name).strip())

        if gastos_sheet_count > 0:
            gastos_breakdown.append((fpath.split('/')[-1], sh, gastos_sheet_count, gastos_sheet_total))

    wb.close()


print('\n# REPORTE ANALISIS EXCELS BARTOLO\n')
print('## A — Por catálogo Munify\n')

all_sheets = set()
for v in sheets_per_file.values():
    all_sheets.update(v)
deps = set()
for s in all_sheets:
    sn = s.strip().lower()
    sn = ''.join(c for c in unicodedata.normalize('NFKD', sn) if not unicodedata.combining(c))
    if 'concejal' in sn:
        deps.add('Concejo Deliberante')
    elif 'turismo' in sn or 'cultura' in sn:
        deps.add('Turismo y Cultura')
    elif 'prof' in sn and 'pub' in sn:
        deps.add('Profesionales y Publicidad')
    elif 'ayuda social' in sn:
        deps.add('Desarrollo Social / Ayuda Social')
    elif 'gerontol' in sn:
        deps.add('Apoyo Gerontológico')
    elif 'empleados' in sn:
        deps.add('Personal Municipal (Empleados de planta)')
    elif 'vestuario' in sn:
        deps.add('Vestuarios')
deps.add('Tesorería / Gastos generales')
print(f'**1. Dependencias / Secretarías inferidas:** {len(deps)}')
for d in sorted(deps):
    print(f'   - {d}')

print()
total_contacts = set()
for cat, names in contacts_by_cat.items():
    total_contacts |= names
print(f'**2. Contactos (personas) únicos deduplicados:** {len(total_contacts)} total\n')
print('   Por categoría inferida:')
for cat in ['Empleados de planta', 'Concejales', 'Personal de Turismo y cultura',
            'Profesionales y publicidad', 'Ayuda social', 'Apoyo gerontológico',
            'Proveedores / contratistas']:
    print(f'   - {cat}: {len(contacts_by_cat.get(cat, set()))}')

print()
print(f'**3. Proyectos / obras únicos inferidos:** {len(proyectos)}')
for p in sorted(proyectos)[:30]:
    print(f'   - {p}')
if len(proyectos) > 30:
    print(f'   ... ({len(proyectos) - 30} mas)')

print()
print(f'**4. Cajas / fondos (variantes de Modo):** {len(cajas)} variantes únicas\n')
print('   | Modo | Conteo | Monto total |')
print('   |---|---|---|')
for m, c in cajas.most_common():
    print(f'   | {m} | {c} | {cajas_monto[m]:,.0f} |')

print('\n## B — Gastos (movimientos de dinero)\n')
print(f'**5. Cantidad total de gastos (filas con monto>0):** {gastos_count}\n')
print('**6. Breakdown por archivo y hoja:**\n')
print('   | Archivo | Hoja | Cant gastos | Monto total |')
print('   |---|---|---|---|')
for fn, sh, c, t in gastos_breakdown:
    print(f'   | {fn} | {sh} | {c} | {t:,.0f} |')

print(f'\n**7. Breakdown por año:**')
for y, c in sorted(gastos_por_anio.items()):
    print(f'   - {y}: {c}')

print('\n## C — Tipos de empleado / subtipos\n')
print(f'**8. Empleados Blanco/Negro:**')
for k, v in empleados_blanco_negro.items():
    print(f'   - {k}: {v}')
print(f'\n**9. Modalidad de pago detectada (Empleados):**')
for k, v in modalidad_pago.items():
    print(f'   - {k}: {v}')

print('\n## D — Conceptos / verbos de gasto\n')
print('**10. Mapeo automático de gastos a los 32 conceptos oficiales:**\n')
print('   | Concepto | Cantidad |')
print('   |---|---|')
for c in CONCEPTS.keys():
    print(f'   | {c} | {concept_counts.get(c, 0)} |')

print(f'\n**Suma de mapeos:** {sum(concept_counts.values())} (debe coincidir con total gastos: {gastos_count})')
