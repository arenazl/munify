"""Importador exhaustivo de los 4 Excel de Bartolo (San Pedro Norte).

Reemplaza los contactos DEMO por reales extraidos del Excel, agrega los
que faltan, y opcionalmente importa gastos historicos.

Fuente:
  - docs/bartolo/Control.xlsx                            (2025, hojas mensuales)
  - docs/bartolo/Gatos municipalidad oct-nov-dic.xlsx    (2025 Q4)
  - docs/bartolo/Gastos mensuales 2025 - 01-09.xlsx      (2025 misc)
  - docs/bartolo/Gastos 2024.xlsx                        (2024 matriz)

Marca:
  - Contactos importados: notas = "[BARTOLO] Importado de Excel <archivo>"
  - Gastos importados: descripcion = "[BARTOLO] <sheet>"
"""
import asyncio
import re
import sys
from datetime import date
from decimal import Decimal
from pathlib import Path
import openpyxl
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent.parent))

from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy import select, update

from models import (
    Municipio, Contacto, Gasto, GastoCuota, User, TipoContacto,
    DestinoGasto, TipoFinanciacion, FormaPago, EstadoGastoCuota,
)
from core.config import settings

BASE = Path('D:/Code/sugerenciasMun/docs/bartolo')
SPN = 'san-pedro-norte'
MESES_NOMBRE_TO_NUM = {
    'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4, 'Mayo': 5, 'Junio': 6,
    'Julio': 7, 'Agosto': 8, 'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12,
}


def normalize_name(s: str) -> str:
    """Limpieza basica para comparar nombres."""
    s = re.sub(r'\s+', ' ', s.strip())
    s = s.replace("''", '').replace('"', '')
    return s


def clasificar(nombre: str) -> tuple[TipoContacto, str | None]:
    """Heuristica para detectar tipo de contacto + subtipo del nombre."""
    n = nombre.lower()

    # Electricista / plomero / albañil
    if 'electric' in n: return (TipoContacto.CONTRATISTA, 'Electricista')
    if 'plomer' in n or 'gasist' in n: return (TipoContacto.CONTRATISTA, 'Plomero')
    if 'albañil' in n or 'mason' in n: return (TipoContacto.CONTRATISTA, 'Albañil')

    # Profesionales (titulo)
    if any(t in n for t in [' abogado', 'abogad', 'dr ', 'dra ', 'doctor', 'dr.', 'medico', 'médico']):
        return (TipoContacto.PROFESIONAL, 'Abogado/Médico')
    if 'contador' in n or 'contad' in n: return (TipoContacto.PROFESIONAL, 'Contador')
    if 'arquitect' in n: return (TipoContacto.PROFESIONAL, 'Arquitecto')
    if 'ingenier' in n: return (TipoContacto.PROFESIONAL, 'Ingeniero')
    if 'nutricion' in n: return (TipoContacto.PROFESIONAL, 'Nutricionista')
    if 'param' in n: return (TipoContacto.PROFESIONAL, 'Paramédico')
    if 'higiene' in n or 'higuiene' in n or 'seguridad' in n: return (TipoContacto.PROFESIONAL, 'Higiene y Seguridad')
    if 'asistente social' in n: return (TipoContacto.PROFESIONAL, 'Asistente social')
    if 'escribano' in n or 'escribana' in n: return (TipoContacto.PROFESIONAL, 'Escribano')

    # Medios y publicidad
    if any(t in n for t in ['canal', 'radio', ' fm ', 'fm ', 'actualidad', 'noticias', 'mir', 'estación', 'estacion ', 'tv ']):
        return (TipoContacto.PROVEEDOR, 'Medios / publicidad')
    if 'directv' in n or 'claro' in n or 'starlink' in n or 'movistar' in n:
        return (TipoContacto.PROVEEDOR, 'Telecomunicaciones')

    # Combustibles
    if 'ypf' in n or 'shell' in n or 'axion' in n or 'puma ' in n: return (TipoContacto.PROVEEDOR, 'Combustibles')

    # Comercios/insumos
    if 'panaderia' in n or 'panadería' in n: return (TipoContacto.PROVEEDOR, 'Panadería')
    if 'imprenta' in n or 'impacto color' in n: return (TipoContacto.PROVEEDOR, 'Imprenta')
    if 'corral' in n or 'acerco' in n: return (TipoContacto.PROVEEDOR, 'Corralón')
    if 'ferret' in n: return (TipoContacto.PROVEEDOR, 'Ferretería')
    if 'agro' in n: return (TipoContacto.PROVEEDOR, 'Agroinsumos')
    if 'pc solutions' in n or 'sistema' in n.split(): return (TipoContacto.PROVEEDOR, 'Sistemas')
    if 'sonido' in n: return (TipoContacto.PROVEEDOR, 'Sonido / eventos')

    # Hoteles / viajes
    if 'hotel' in n: return (TipoContacto.PROVEEDOR, 'Hotel')
    if 'remis' in n or 'transporte' in n or 'transport' in n: return (TipoContacto.PROVEEDOR, 'Transporte')

    # Seguros
    if 'seguro' in n or 'previnca' in n: return (TipoContacto.PROVEEDOR, 'Seguros')

    # Deportes / entrenadores
    if any(t in n for t in ['boxeo', 'deporte', 'entrenador', 'entrenam', 'arquero', 'fútbol', 'futbol', 'liga']):
        return (TipoContacto.CONTRATISTA, 'Entrenador deportivo')
    if 'cultura' in n or 'arte' in n or 'taller' in n or 'folcklor' in n or 'folklor' in n:
        return (TipoContacto.PROVEEDOR, 'Cultura')

    # Ayuda social / asociaciones
    if 'ayuda social' in n: return (TipoContacto.BENEFICIARIO, 'Ayuda social')
    if 'gerontol' in n: return (TipoContacto.BENEFICIARIO, 'Apoyo gerontológico')
    if 'club ' in n or 'cooperat' in n or 'asoc' in n or 'comisi' in n:
        return (TipoContacto.BENEFICIARIO, 'Asociación / club')

    # Cooperativas / servicios públicos
    if 'consorcio' in n or 'caminero' in n: return (TipoContacto.PROVEEDOR, 'Servicios')

    # Por defecto: proveedor (la mayoria de empresa en hojas mensuales son proveedores)
    return (TipoContacto.PROVEEDOR, None)


def alias_pago(nombre: str) -> str:
    """Genera alias de pago consistente."""
    parts = re.findall(r'[A-Za-zÁ-ÿ]+', nombre)
    return '.'.join(p.upper() for p in parts[:3])[:60]


MAPEO_MODO = {
    'transf': FormaPago.TRANSFERENCIA,
    'transferencia': FormaPago.TRANSFERENCIA,
    'trans': FormaPago.TRANSFERENCIA,
    'efectivo': FormaPago.EFECTIVO,
    'eftv': FormaPago.EFECTIVO,
    'cheque': FormaPago.CHEQUE,
    'echeq': FormaPago.CHEQUE,
    'visa': FormaPago.TARJETA,
    'mastercard': FormaPago.TARJETA,
    'tarjeta': FormaPago.TARJETA,
    'debin': FormaPago.TRANSFERENCIA,
    'vep': FormaPago.TRANSFERENCIA,
}


def mapear_modo(modo_str: str | None) -> FormaPago:
    if not modo_str:
        return FormaPago.OTRO
    s = str(modo_str).lower()
    for k, v in MAPEO_MODO.items():
        if k in s:
            return v
    return FormaPago.OTRO


def es_modo_valido(modo_str: str | None) -> bool:
    """Detecta si la celda 'Modo' es realmente un metodo de pago, no
    basura como una fecha o un nombre repetido."""
    if not modo_str:
        return True
    s = str(modo_str).lower()
    # Si es una fecha (contiene /) probablemente es un error de carga
    if '/' in s and len(s) > 5:
        return False
    return True


# Mapeo archivo + hoja → (año, mes opcional)
# Asumimos:
# - Control.xlsx monthly sheets → 2025 (Bartolo nos paso esto como su workbook actual)
# - Gatos oct-nov-dic.xlsx → 2025 Q4
# - Gastos mensuales 2025 - 01-09.xlsx Octubre → 2025-10 (duplicado, skip)
# - Gastos 2024.xlsx → 2024, formato matriz (complejo, skip por ahora)
FUENTES_GASTOS = [
    # (filename, sheet_name, año, mes)
    ('Control.xlsx', 'Enero', 2025, 1),
    ('Control.xlsx', 'Febrero', 2025, 2),
    ('Control.xlsx', 'Marzo', 2025, 3),
    ('Control.xlsx', 'Abril', 2025, 4),
    # Mayo-Diciembre estan vacios en Control
    ('Gatos municipalidad oct-nov-dic.xlsx', 'Octubre', 2025, 10),
    ('Gatos municipalidad oct-nov-dic.xlsx', 'Noviembre', 2025, 11),
    ('Gatos municipalidad oct-nov-dic.xlsx', 'Diciembre', 2025, 12),
]


async def main():
    engine = create_async_engine(settings.DATABASE_URL)
    SessionLocal = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with SessionLocal() as db:
        # ==================== Setup ====================
        muni = (await db.execute(select(Municipio).where(Municipio.codigo == SPN))).scalar_one_or_none()
        if not muni:
            print(f"[!] muni '{SPN}' no existe"); return
        mid = muni.id

        admin = (await db.execute(
            select(User).where(User.email == 'admin@san-pedro-norte.demo.com')
        )).scalar_one_or_none()
        if not admin:
            print("[!] admin no existe"); return

        # ==================== Phase 1: Desactivar DEMOs anteriores ====================
        print("=" * 60)
        print("PHASE 1: Desactivar contactos [DEMO] previos")
        print("=" * 60)
        r = await db.execute(
            update(Contacto)
            .where(Contacto.municipio_id == mid)
            .where(Contacto.notas.like('%[DEMO]%'))
            .values(activo=False)
        )
        print(f"  Desactivados: {r.rowcount} contactos DEMO")
        await db.commit()

        # ==================== Phase 2: Extract real contacts ====================
        print("\n" + "=" * 60)
        print("PHASE 2: Extraer contactos reales del Excel")
        print("=" * 60)

        contactos_reales: dict[str, dict] = {}  # nombre_normalizado -> { nombre, freq, archivos }

        for fname in os.listdir(BASE):
            if not fname.endswith('.xlsx'): continue
            p = BASE / fname
            wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), (None,))
                if not first or first[0] != 'Monto':
                    continue
                for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
                    monto = row[0]; emp = row[1] if len(row) > 1 else None
                    if isinstance(monto, (int, float)) and monto > 0 and emp:
                        nombre = normalize_name(str(emp))
                        if len(nombre) < 2 or nombre == 'None': continue
                        # Saltear nombres que son numeros, montos
                        if nombre.replace('.', '').replace(',', '').replace(' ', '').isdigit(): continue
                        key = nombre.lower()
                        if key in contactos_reales:
                            contactos_reales[key]['freq'] += 1
                        else:
                            contactos_reales[key] = {'nombre': nombre, 'freq': 1, 'archivo': fname}
            wb.close()

        print(f"  {len(contactos_reales)} nombres distintos encontrados")

        # ==================== Phase 3: Insertar contactos reales ====================
        print("\n" + "=" * 60)
        print("PHASE 3: Insertar contactos reales (skipping duplicados activos)")
        print("=" * 60)

        # Cargar todos los contactos activos existentes para no duplicar
        existing = (await db.execute(
            select(Contacto).where(Contacto.municipio_id == mid, Contacto.activo == True)  # noqa: E712
        )).scalars().all()
        existing_names = {(c.nombre + ' ' + (c.apellido or '')).strip().lower(): c.id for c in existing}

        creados = 0
        skipped = 0
        contacto_por_nombre: dict[str, int] = {}  # mapping para gastos

        for key, data in contactos_reales.items():
            nombre = data['nombre']
            nombre_lower = nombre.lower()
            if nombre_lower in existing_names:
                contacto_por_nombre[nombre_lower] = existing_names[nombre_lower]
                skipped += 1
                continue

            tipo, subtipo = clasificar(nombre)
            # Heuristica: si tiene 2-3 palabras parece persona, sino empresa
            partes = nombre.split()
            if len(partes) >= 2 and not any(t in nombre.lower() for t in ['canal', 'radio', 'panaderia', 'ferret', 'corralon', 'agro', 'pc solutions', 'mu patronal', 'la boutique', 'fm', 'sonido']):
                nombre_corto = partes[0]
                apellido = ' '.join(partes[1:])[:100]
            else:
                nombre_corto = nombre[:100]
                apellido = None

            c = Contacto(
                municipio_id=mid,
                nombre=nombre_corto,
                apellido=apellido,
                tipo=tipo,
                subtipo=subtipo,
                alias_pago=alias_pago(nombre),
                notas=f"[BARTOLO] Importado de Excel ({data['archivo']}) - {data['freq']} apariciones",
                activo=True,
            )
            db.add(c)
            creados += 1

        await db.flush()
        # Re-leer para llenar el mapa
        all_now = (await db.execute(
            select(Contacto).where(Contacto.municipio_id == mid, Contacto.activo == True)  # noqa: E712
        )).scalars().all()
        for c in all_now:
            full = (c.nombre + ' ' + (c.apellido or '')).strip().lower()
            contacto_por_nombre[full] = c.id
            # tambien la version "solo nombre"
            contacto_por_nombre[c.nombre.lower()] = c.id

        await db.commit()
        print(f"  Creados: {creados}, Saltados (ya existian): {skipped}")

        # ==================== Phase 4: Importar gastos historicos ====================
        print("\n" + "=" * 60)
        print("PHASE 4: Importar gastos historicos")
        print("=" * 60)

        # Borrar gastos previos importados de [BARTOLO] para evitar duplicados
        prev = (await db.execute(
            select(Gasto.id).where(
                Gasto.municipio_id == mid,
                Gasto.descripcion.like('%[BARTOLO]%'),
            )
        )).scalars().all()
        if prev:
            print(f"  Borrando {len(prev)} gastos [BARTOLO] previos...")
            # Borrar cuotas asociadas primero
            from sqlalchemy import delete
            await db.execute(delete(GastoCuota).where(GastoCuota.gasto_id.in_(prev)))
            await db.execute(delete(Gasto).where(Gasto.id.in_(prev)))
            await db.commit()

        gastos_creados = 0
        gastos_skip = 0
        for fname, sname, anio, mes in FUENTES_GASTOS:
            p = BASE / fname
            if not p.exists(): continue
            wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
            if sname not in wb.sheetnames:
                wb.close(); continue
            ws = wb[sname]
            fecha_gasto = date(anio, mes, 15)  # mid-mes como aproximacion

            for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
                monto, emp, modo = row[0], (row[1] if len(row) > 1 else None), (row[2] if len(row) > 2 else None)
                if not isinstance(monto, (int, float)) or monto <= 0: continue
                if not emp: continue
                nombre = normalize_name(str(emp))
                if len(nombre) < 2 or nombre.replace('.', '').replace(',', '').replace(' ', '').isdigit():
                    continue

                cid = contacto_por_nombre.get(nombre.lower())
                if not cid:
                    # Buscar por primer palabra del nombre
                    primera = nombre.split()[0].lower() if nombre.split() else ''
                    for k, v in contacto_por_nombre.items():
                        if k.startswith(primera) and len(primera) > 3:
                            cid = v; break
                if not cid:
                    gastos_skip += 1; continue

                forma = mapear_modo(modo) if es_modo_valido(modo) else FormaPago.OTRO
                # Concepto: usar la sheet o fallback
                concepto = sname if sname not in MESES_NOMBRE_TO_NUM else 'Pago'

                g = Gasto(
                    municipio_id=mid,
                    creador_id=admin.id,
                    destino_tipo=DestinoGasto.CONTACTO,
                    destino_contacto_id=cid,
                    destino_dependencia_id=None,
                    concepto=concepto,
                    descripcion=f"[BARTOLO] {fname} / {sname}",
                    monto_pesos=Decimal(str(monto)),
                    fecha=fecha_gasto,
                    tipo_financiacion=TipoFinanciacion.CONTADO,
                    forma_pago=forma,
                    activo=True,
                )
                db.add(g)
                await db.flush()
                # Cuota unica pagada
                db.add(GastoCuota(
                    gasto_id=g.id, numero=1, monto=Decimal(str(monto)),
                    fecha_vencimiento=fecha_gasto, fecha_pago=fecha_gasto,
                    estado=EstadoGastoCuota.PAGADA, forma_pago=forma,
                ))
                gastos_creados += 1
            wb.close()

        await db.commit()
        print(f"  Gastos creados: {gastos_creados}, Skip (sin contacto match): {gastos_skip}")

        # ==================== Phase 5: Resumen ====================
        print("\n" + "=" * 60)
        print("RESUMEN FINAL")
        print("=" * 60)

        from sqlalchemy import func as sa_func
        # Contactos por estado
        total_act = (await db.execute(
            select(sa_func.count(Contacto.id)).where(Contacto.municipio_id == mid, Contacto.activo == True)  # noqa: E712
        )).scalar()
        total_inact = (await db.execute(
            select(sa_func.count(Contacto.id)).where(Contacto.municipio_id == mid, Contacto.activo == False)  # noqa: E712
        )).scalar()
        bartolo_count = (await db.execute(
            select(sa_func.count(Contacto.id)).where(
                Contacto.municipio_id == mid, Contacto.notas.like('%[BARTOLO]%'),
            )
        )).scalar()
        demo_count = (await db.execute(
            select(sa_func.count(Contacto.id)).where(
                Contacto.municipio_id == mid, Contacto.notas.like('%[DEMO]%'),
            )
        )).scalar()

        print(f"\nContactos:")
        print(f"  Activos:       {total_act}")
        print(f"  Inactivos:     {total_inact}")
        print(f"  De Bartolo:    {bartolo_count}")
        print(f"  Marcados DEMO: {demo_count}")
        print(f"  REALES sin marcar (cargados en seed honesto): {total_act - bartolo_count - demo_count}")

        # Gastos
        gastos_bartolo = (await db.execute(
            select(sa_func.count(Gasto.id)).where(
                Gasto.municipio_id == mid, Gasto.descripcion.like('%[BARTOLO]%'),
            )
        )).scalar()
        gastos_total = (await db.execute(
            select(sa_func.count(Gasto.id)).where(Gasto.municipio_id == mid, Gasto.activo == True)  # noqa: E712
        )).scalar()
        print(f"\nGastos:")
        print(f"  Total activos:    {gastos_total}")
        print(f"  De Bartolo:       {gastos_bartolo}")
        print(f"  Otros (sueldos manuales del seed honesto): {gastos_total - gastos_bartolo}")

    await engine.dispose()


if __name__ == '__main__':
    import os
    asyncio.run(main())
