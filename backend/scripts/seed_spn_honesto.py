"""Seed HONESTO de San Pedro Norte.

Estrategia:
  1. Lee los datos REALES del Excel `docs/bartolo/Control.xlsx` y los carga
     tal cual estan (concejales con sus dietas reales, empleados con sueldos
     base reales, turismo y cultura, gastos sueltos de los meses).
  2. Agrega contactos/gastos DUMMY marcados con prefijo `[DEMO]` en
     `notas` para que el intendente pueda mostrar features (prestamos a
     cuotas, profesionales con honorarios, etc.). El marcador es visible
     para que nunca se confunda con data real.
  3. Asigna alias_pago a TODOS: los que tienen alias real del Excel se
     respetan, los demas reciben uno generado tipo `NOMBRE.APELLIDO.spn`
     (consistente para la demo).
  4. Restaura las coords previas del backup (las que quedaron bien
     dispersas en el mapa). Los contactos nuevos toman coords del anillo
     0.5-5km del centro de SPN.

Idempotente: borra todo lo del muni 80 antes de cargar.

Uso:
    cd backend && python scripts/seed_spn_honesto.py
"""
import asyncio
import json
import math
import random
import sys
from datetime import date
from decimal import Decimal
from pathlib import Path
from calendar import monthrange

ROOT = Path(__file__).resolve().parent.parent
PROJECT_ROOT = ROOT.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker

from core.config import settings
from models import (
    Municipio, MunicipioDependencia, MunicipioModulo,
    Dependencia, Contacto, Gasto, GastoCuota, User,
    TipoContacto, DestinoGasto, TipoFinanciacion, FrecuenciaRecurrencia,
    FormaPago, EstadoGastoCuota, RolUsuario,
)

EXCEL_PATH = PROJECT_ROOT / 'docs' / 'bartolo' / 'Control.xlsx'
COORDS_BACKUP = Path('/tmp/coords_backup.json')

SPN_CODIGO = 'san-pedro-norte'
SPN_LAT = -30.266
SPN_LON = -64.125
MIN_KM = 0.5
MAX_KM = 5.0

random.seed(42)


def random_punto_anillo():
    ang = random.uniform(0, 2 * math.pi)
    d = random.uniform(MIN_KM, MAX_KM)
    dlat = (d / 111.0) * math.sin(ang)
    dlon = (d / (111.0 * math.cos(math.radians(SPN_LAT)))) * math.cos(ang)
    return SPN_LAT + dlat, SPN_LON + dlon


def gen_alias(nombre: str, apellido: str | None) -> str:
    """Genera alias dummy consistente, marcado como demo via convencion."""
    parts = [nombre]
    if apellido:
        parts.append(apellido.split()[0])
    parts.append('spn')
    return '.'.join(p.upper().replace(' ', '') for p in parts)


def add_months(d: date, n: int) -> date:
    total = d.month - 1 + n
    y = d.year + total // 12
    m = total % 12 + 1
    last = monthrange(y, m)[1]
    return date(y, m, min(d.day, last))


MES_A_NUM = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
    'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
    'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
}


# ============================================================
# 1. Leer Excel REAL
# ============================================================

def leer_excel():
    """Devuelve dict con datos del Excel del intendente."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    out = {
        'concejales': [],    # lista de {nombre, apellido, alias, gastos: [{mes, monto}]}
        'empleados': [],     # idem (sueldo base en sueldo_base)
        'turismo_cultura': [],
        'mensuales': [],     # gastos sueltos sin destinatario fijo: {mes, monto, descripcion}
    }

    # Sheet "Concejales"
    if 'Concejales' in wb.sheetnames:
        ws = wb['Concejales']
        header = list(next(ws.iter_rows(values_only=True), []))
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            nombre_full = str(row[0]).strip()
            if not nombre_full or nombre_full.lower().startswith(('gastos', 'total', 'sub')):
                continue
            partes = nombre_full.split(' ', 1)
            nombre = partes[0]
            apellido = partes[1] if len(partes) > 1 else None
            alias = row[1] if len(row) > 1 and row[1] and isinstance(row[1], str) else None
            # Columnas con montos: idx 2=Enero, 4=Febrero, etc. Veamos header.
            gastos = []
            for idx, val in enumerate(row):
                if not isinstance(val, (int, float)) or val <= 0:
                    continue
                col_name = header[idx] if idx < len(header) else None
                if not col_name:
                    continue
                key = str(col_name).lower().strip()
                mes_num = MES_A_NUM.get(key)
                concepto = None
                if mes_num:
                    concepto = f'Dieta {key.capitalize()}'
                elif 'aguinaldo' in key and 'medio' in key:
                    concepto = 'Medio aguinaldo'
                    mes_num = 6
                elif 'aguinaldo' in key:
                    concepto = 'Aguinaldo'
                    mes_num = 12
                if concepto and mes_num:
                    gastos.append({'mes': mes_num, 'monto': Decimal(str(val)), 'concepto': concepto})
            out['concejales'].append({
                'nombre': nombre, 'apellido': apellido, 'alias': alias, 'gastos': gastos,
            })

    # Sheet "Empleados"
    if 'Empleados' in wb.sheetnames:
        ws = wb['Empleados']
        # Estructura: col 0=numero, col 1=sueldo_base, col 2=nombre completo, col 3=alias
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3 or not row[2]:
                continue
            sueldo = row[1] if isinstance(row[1], (int, float)) else None
            nombre_full = str(row[2]).strip()
            if not nombre_full or nombre_full.lower().startswith(('total', 'gastos')):
                continue
            partes = nombre_full.split(' ', 1)
            # Orden en Excel: "Apellido Nombre" (ej "Arias Pilar")
            if len(partes) >= 2:
                apellido = partes[0]
                nombre = ' '.join(partes[1:])
            else:
                nombre = nombre_full
                apellido = None
            alias = row[3] if len(row) > 3 and isinstance(row[3], str) and row[3].strip() else None
            out['empleados'].append({
                'nombre': nombre, 'apellido': apellido, 'alias': alias,
                'sueldo_base': Decimal(str(sueldo)) if sueldo and sueldo > 0 else None,
            })

    # Sheet "Turismo y cultura"
    if 'Turismo y cultura' in wb.sheetnames:
        ws = wb['Turismo y cultura']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            nombre_full = str(row[0]).strip()
            if not nombre_full or nombre_full.lower().startswith(('total', 'gastos')):
                continue
            partes = nombre_full.split(' ', 1)
            nombre = partes[0]
            apellido = partes[1] if len(partes) > 1 else None
            out['turismo_cultura'].append({'nombre': nombre, 'apellido': apellido})

    # Sheets mensuales (gastos sueltos)
    for sn in ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']:
        if sn not in wb.sheetnames:
            continue
        mes_num = MES_A_NUM[sn.lower()]
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            monto = None
            desc = None
            for v in row:
                if isinstance(v, (int, float)) and v > 0 and monto is None:
                    monto = Decimal(str(v))
                elif isinstance(v, str) and v.strip() and desc is None:
                    desc = v.strip()
            if monto and desc:
                out['mensuales'].append({'mes': mes_num, 'monto': monto, 'descripcion': desc})

    return out


# ============================================================
# Dummies MARCADOS [DEMO]
# ============================================================

DUMMY_PROFESIONALES = [
    ('Roberto', 'Méndez', 'abogado'),
    ('Susana', 'Pérez', 'contador'),
    ('Hugo', 'Castro', 'doctor'),
]

DUMMY_BENEFICIARIOS = [
    ('Juan', 'González', 'Préstamo agrario', 800000, 6),
    ('Marta', 'Fernández', 'Préstamo productivo', 500000, 12),
    ('Ramón', 'Ledesma', 'Préstamo personal', 300000, 6),
]


# ============================================================
# Main
# ============================================================

async def main():
    print(f'Leyendo Excel: {EXCEL_PATH}')
    data = leer_excel()
    print(f'  - {len(data["concejales"])} concejales del Excel')
    print(f'  - {len(data["empleados"])} empleados del Excel')
    print(f'  - {len(data["turismo_cultura"])} turismo y cultura del Excel')
    print(f'  - {len(data["mensuales"])} gastos sueltos en sheets mensuales')

    # Cargar backup coords
    coords_backup = {}
    if COORDS_BACKUP.exists():
        with open(COORDS_BACKUP, encoding='utf-8') as f:
            for item in json.load(f):
                key = f"{item['nombre']}|{item.get('apellido') or ''}".lower()
                coords_backup[key] = (item['lat'], item['lon'])
        print(f'  - {len(coords_backup)} coords del backup')

    engine = create_async_engine(settings.DATABASE_URL, echo=False)
    SessionLocal = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with SessionLocal() as session:
        muni = (await session.execute(select(Municipio).where(Municipio.codigo == SPN_CODIGO))).scalar_one()
        admin = (await session.execute(
            select(User).where(User.municipio_id == muni.id, User.rol == RolUsuario.ADMIN)
        )).scalar_one()
        print(f'\nMuni: {muni.nombre} (id={muni.id}), admin={admin.email}')

        # ---- Limpiar ----
        await session.execute(text("""
            DELETE FROM gastos_cuotas WHERE gasto_id IN (
                SELECT id FROM gastos WHERE municipio_id = :mid
            )
        """), {'mid': muni.id})
        await session.execute(text('DELETE FROM gastos WHERE municipio_id = :mid'), {'mid': muni.id})
        await session.execute(text('DELETE FROM contactos WHERE municipio_id = :mid'), {'mid': muni.id})
        await session.commit()
        print('Wipe completo del muni')

        def coord_for(nombre, apellido):
            key = f"{nombre}|{apellido or ''}".lower()
            if key in coords_backup:
                return coords_backup[key]
            return random_punto_anillo()

        # ---- Crear concejales ----
        contactos_creados = []
        for c in data['concejales']:
            lat, lon = coord_for(c['nombre'], c['apellido'])
            ct = Contacto(
                municipio_id=muni.id,
                nombre=c['nombre'], apellido=c['apellido'],
                alias_pago=c['alias'] or gen_alias(c['nombre'], c['apellido']),
                tipo=TipoContacto.CONCEJAL,
                latitud=lat, longitud=lon,
                notas=None if c['alias'] else '[alias generado para demo]',
            )
            session.add(ct)
            contactos_creados.append((ct, c.get('gastos', [])))

        # Eduardo García (KMZ) - lo agrego como CONCEJAL o BENEFICIARIO?
        # En el KMZ es solo un placemark sin tipo. Por ahora BENEFICIARIO.
        eduardo = Contacto(
            municipio_id=muni.id,
            nombre='Eduardo', apellido='García',
            alias_pago=gen_alias('Eduardo', 'García'),
            tipo=TipoContacto.BENEFICIARIO,
            latitud=-30.26586490649007, longitud=-64.12454421138095,
            notas='[ubicación real del KMZ]',
        )
        session.add(eduardo)
        contactos_creados.append((eduardo, []))

        # ---- Empleados ----
        empleados_contactos = []
        for e in data['empleados']:
            lat, lon = coord_for(e['nombre'], e['apellido'])
            ct = Contacto(
                municipio_id=muni.id,
                nombre=e['nombre'], apellido=e['apellido'],
                alias_pago=e['alias'] or gen_alias(e['nombre'], e['apellido']),
                tipo=TipoContacto.EMPLEADO,
                latitud=lat, longitud=lon,
                notas=None if e['alias'] else '[alias generado para demo]',
            )
            session.add(ct)
            empleados_contactos.append((ct, e.get('sueldo_base')))

        # ---- Turismo y cultura ----
        for t in data['turismo_cultura']:
            lat, lon = coord_for(t['nombre'], t['apellido'])
            ct = Contacto(
                municipio_id=muni.id,
                nombre=t['nombre'], apellido=t['apellido'],
                alias_pago=gen_alias(t['nombre'], t['apellido']),
                tipo=TipoContacto.BENEFICIARIO,
                subtipo='turismo y cultura',
                latitud=lat, longitud=lon,
                notas='[alias generado para demo]',
            )
            session.add(ct)

        # ---- Gastos sueltos REALES de meses ----
        # Crear contactos especiales para los gastos sueltos
        extras_real = {}  # nombre -> contacto
        for g in data['mensuales']:
            desc = g['descripcion']
            # Resolver nombre / apellido
            if 'fofindes' in desc.lower():
                continue  # Es empresa, lo manejamos como gasto a dependencia abajo
            if 'mercedes' in desc.lower():
                nombre_clave = 'Mercedes Villarroel'
                nombre = 'Mercedes'; apellido = 'Villarroel'
            elif 'vanesa' in desc.lower():
                nombre_clave = 'Vanesa Suárez'
                nombre = 'Vanesa'; apellido = 'Suárez'
            elif 'negro de oro' in desc.lower():
                nombre_clave = 'Negro de Oro'
                nombre = 'Negro'; apellido = 'de Oro'
            else:
                nombre_clave = desc[:40]
                partes = desc.split(' ', 1)
                nombre = partes[0]
                apellido = partes[1] if len(partes) > 1 else None

            if nombre_clave not in extras_real:
                lat, lon = coord_for(nombre, apellido)
                ct = Contacto(
                    municipio_id=muni.id,
                    nombre=nombre, apellido=apellido,
                    alias_pago=gen_alias(nombre, apellido),
                    tipo=TipoContacto.BENEFICIARIO,
                    latitud=lat, longitud=lon,
                    notas=f'[del sheet mensual del Excel]',
                )
                session.add(ct)
                await session.flush()
                extras_real[nombre_clave] = ct
            else:
                ct = extras_real[nombre_clave]

        # ---- Dummies MARCADOS [DEMO] ----
        dummies_prof = []
        for n, a, sub in DUMMY_PROFESIONALES:
            lat, lon = coord_for(n, a)
            ct = Contacto(
                municipio_id=muni.id,
                nombre=n, apellido=a,
                alias_pago=gen_alias(n, a),
                tipo=TipoContacto.PROFESIONAL, subtipo=sub,
                latitud=lat, longitud=lon,
                notas='[DEMO] contacto generado para mostrar feature',
            )
            session.add(ct)
            dummies_prof.append(ct)

        dummies_benef = []
        for n, a, concepto, monto, cuotas in DUMMY_BENEFICIARIOS:
            lat, lon = coord_for(n, a)
            ct = Contacto(
                municipio_id=muni.id,
                nombre=n, apellido=a,
                alias_pago=gen_alias(n, a),
                tipo=TipoContacto.BENEFICIARIO,
                latitud=lat, longitud=lon,
                notas='[DEMO] contacto generado para mostrar feature',
            )
            session.add(ct)
            dummies_benef.append((ct, concepto, Decimal(str(monto)), cuotas))

        await session.flush()
        print(f'Contactos creados (reales + dummies)')

        # ============================================================
        # GASTOS
        # ============================================================
        anio = 2026
        gastos_count = 0

        # Concejales: gastos REALES del Excel
        for ct, gastos in contactos_creados:
            for g in gastos:
                fecha = date(anio, g['mes'], 1)
                gasto = Gasto(
                    municipio_id=muni.id, creador_id=admin.id,
                    destino_tipo=DestinoGasto.CONTACTO, destino_contacto_id=ct.id,
                    concepto=g['concepto'], monto_pesos=g['monto'], fecha=fecha,
                    tipo_financiacion=TipoFinanciacion.CONTADO,
                    forma_pago=FormaPago.TRANSFERENCIA,
                )
                session.add(gasto)
                await session.flush()
                session.add(GastoCuota(
                    gasto_id=gasto.id, numero=1, monto=g['monto'],
                    fecha_vencimiento=fecha, fecha_pago=fecha,
                    estado=EstadoGastoCuota.PAGADA, forma_pago=FormaPago.TRANSFERENCIA,
                ))
                gastos_count += 1

        # Empleados: SUELDO BASE como recurrente (REAL del Excel)
        for ct, sueldo in empleados_contactos:
            if not sueldo:
                continue
            fecha_ini = date(anio, 1, 1)
            fecha_fin = date(anio, 12, 31)
            gasto = Gasto(
                municipio_id=muni.id, creador_id=admin.id,
                destino_tipo=DestinoGasto.CONTACTO, destino_contacto_id=ct.id,
                concepto='Sueldo mensual', monto_pesos=sueldo, fecha=fecha_ini,
                tipo_financiacion=TipoFinanciacion.RECURRENTE,
                forma_pago=FormaPago.TRANSFERENCIA,
                frecuencia=FrecuenciaRecurrencia.MENSUAL,
                fecha_fin_recurrencia=fecha_fin,
            )
            session.add(gasto)
            await session.flush()
            f = fecha_ini
            numero = 1
            hoy = date.today()
            while f <= fecha_fin and numero <= 12:
                estado = EstadoGastoCuota.PAGADA if f < hoy else EstadoGastoCuota.PENDIENTE
                session.add(GastoCuota(
                    gasto_id=gasto.id, numero=numero, monto=sueldo,
                    fecha_vencimiento=f,
                    fecha_pago=f if estado == EstadoGastoCuota.PAGADA else None,
                    estado=estado,
                    forma_pago=FormaPago.TRANSFERENCIA if estado == EstadoGastoCuota.PAGADA else None,
                ))
                f = add_months(f, 1)
                numero += 1
            gastos_count += 1

        # Gastos sueltos REALES (Mercedes, Vanesa, Negro de Oro)
        for g in data['mensuales']:
            desc = g['descripcion'].lower()
            if 'fofindes' in desc:
                continue
            ct = None
            concepto = g['descripcion']
            if 'mercedes' in desc:
                ct = extras_real.get('Mercedes Villarroel')
                concepto = 'Ayuda social'
            elif 'vanesa' in desc:
                ct = extras_real.get('Vanesa Suárez')
                concepto = g['descripcion']
            elif 'negro de oro' in desc:
                ct = extras_real.get('Negro de Oro')
                concepto = 'Negro de Oro 2-2'
            if not ct:
                continue
            fecha = date(anio, g['mes'], 1)
            gasto = Gasto(
                municipio_id=muni.id, creador_id=admin.id,
                destino_tipo=DestinoGasto.CONTACTO, destino_contacto_id=ct.id,
                concepto=concepto, monto_pesos=g['monto'], fecha=fecha,
                tipo_financiacion=TipoFinanciacion.CONTADO,
                forma_pago=FormaPago.TRANSFERENCIA,
            )
            session.add(gasto)
            await session.flush()
            session.add(GastoCuota(
                gasto_id=gasto.id, numero=1, monto=g['monto'],
                fecha_vencimiento=fecha, fecha_pago=fecha,
                estado=EstadoGastoCuota.PAGADA, forma_pago=FormaPago.TRANSFERENCIA,
            ))
            gastos_count += 1

        # Gasto a Fofindes (empresa) - asignamos a dependencia de Obras
        dep_obras_q = await session.execute(text(
            "SELECT md.id FROM municipio_dependencias md JOIN dependencias d ON md.dependencia_id = d.id "
            "WHERE md.municipio_id = :mid AND d.nombre LIKE '%Obras%' LIMIT 1"
        ), {'mid': muni.id})
        dep_obras_id = dep_obras_q.scalar()
        if dep_obras_id:
            gasto = Gasto(
                municipio_id=muni.id, creador_id=admin.id,
                destino_tipo=DestinoGasto.DEPENDENCIA, destino_dependencia_id=dep_obras_id,
                concepto='Obra Vestuarios - pago a Fofindes',
                descripcion='Pago de obra vestuarios (empresa: Fofindes)',
                monto_pesos=Decimal('8100771.30'), fecha=date(anio, 4, 1),
                tipo_financiacion=TipoFinanciacion.CONTADO,
                forma_pago=FormaPago.TRANSFERENCIA,
            )
            session.add(gasto)
            await session.flush()
            session.add(GastoCuota(
                gasto_id=gasto.id, numero=1, monto=Decimal('8100771.30'),
                fecha_vencimiento=date(anio, 4, 1), fecha_pago=date(anio, 4, 1),
                estado=EstadoGastoCuota.PAGADA, forma_pago=FormaPago.TRANSFERENCIA,
            ))
            gastos_count += 1

        # Dummies: profesionales con honorario MARCADO [DEMO]
        for ct in dummies_prof:
            fecha = date(anio, 5, 1)
            sub = ct.subtipo or ''
            gasto = Gasto(
                municipio_id=muni.id, creador_id=admin.id,
                destino_tipo=DestinoGasto.CONTACTO, destino_contacto_id=ct.id,
                concepto=f'Honorarios {sub}'.strip(),
                descripcion='[DEMO] gasto generado para mostrar feature',
                monto_pesos=Decimal(random.choice(['180000', '250000'])), fecha=fecha,
                tipo_financiacion=TipoFinanciacion.CONTADO,
                forma_pago=FormaPago.TRANSFERENCIA,
            )
            session.add(gasto)
            await session.flush()
            session.add(GastoCuota(
                gasto_id=gasto.id, numero=1, monto=gasto.monto_pesos,
                fecha_vencimiento=fecha, fecha_pago=fecha,
                estado=EstadoGastoCuota.PAGADA, forma_pago=FormaPago.TRANSFERENCIA,
            ))
            gastos_count += 1

        # Dummies: prestamos en cuotas
        for ct, concepto, monto, n_cuotas in dummies_benef:
            fecha_ini = date(anio, 3, 1)
            monto_cuota = (monto / n_cuotas).quantize(Decimal('0.01'))
            gasto = Gasto(
                municipio_id=muni.id, creador_id=admin.id,
                destino_tipo=DestinoGasto.CONTACTO, destino_contacto_id=ct.id,
                concepto=concepto,
                descripcion='[DEMO] préstamo generado para mostrar feature de cuotas',
                monto_pesos=monto, fecha=fecha_ini,
                tipo_financiacion=TipoFinanciacion.PRESTAMO,
                forma_pago=FormaPago.TRANSFERENCIA,
                cuotas_total=n_cuotas,
            )
            session.add(gasto)
            await session.flush()
            hoy = date.today()
            for i in range(n_cuotas):
                f = add_months(fecha_ini, i)
                estado = EstadoGastoCuota.PAGADA if i < 2 and f <= hoy else EstadoGastoCuota.PENDIENTE
                session.add(GastoCuota(
                    gasto_id=gasto.id, numero=i+1, monto=monto_cuota,
                    fecha_vencimiento=f,
                    fecha_pago=f if estado == EstadoGastoCuota.PAGADA else None,
                    estado=estado,
                    forma_pago=FormaPago.TRANSFERENCIA if estado == EstadoGastoCuota.PAGADA else None,
                ))
            gastos_count += 1

        await session.commit()

        # Resumen final
        c_total = await session.execute(text('SELECT COUNT(*) FROM contactos WHERE municipio_id = :mid'), {'mid': muni.id})
        g_total = await session.execute(text('SELECT COUNT(*) FROM gastos WHERE municipio_id = :mid'), {'mid': muni.id})
        c_demo = await session.execute(text("SELECT COUNT(*) FROM contactos WHERE municipio_id = :mid AND notas LIKE '%[DEMO]%'"), {'mid': muni.id})
        g_demo = await session.execute(text("SELECT COUNT(*) FROM gastos WHERE municipio_id = :mid AND descripcion LIKE '%[DEMO]%'"), {'mid': muni.id})

        print()
        print('=== RESUMEN ===')
        print(f'  Contactos totales: {c_total.scalar()}')
        print(f'  Contactos [DEMO]: {c_demo.scalar()}')
        print(f'  Gastos totales: {g_total.scalar()}')
        print(f'  Gastos [DEMO]: {g_demo.scalar()}')
        print()
        print('[OK] Seed honesto completado.')

    await engine.dispose()


if __name__ == '__main__':
    asyncio.run(main())
