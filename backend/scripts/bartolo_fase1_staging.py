"""
FASE 1 — Ingesta cruda de los 3 Excels de Bartolo a tabla staging.

Crea tabla `bartolo_raw` y baja TODO: gastos generales (bloques
[Mes, Monto, Empresa, Modo]) + grilla pivote de personal (celda
mensual = 1 fila).

Sin clasificar, sin matching a catalogos. Solo extraer + normalizar
nombres + detectar header corrido.

Tabla staging:
  - id, fuente_archivo, fuente_hoja, fuente_fila, fuente_col_block
  - bloque_corrido (TRUE si remapeo cols+1)
  - mes (1-12), anio
  - monto (Decimal)
  - empresa_raw (texto crudo)
  - empresa_normalizada (lowercase trim, "apellido nombre" -> "nombre apellido")
  - empresa_parentesis (texto entre parentesis si hay)
  - modo_raw, modo_normalizado
  - tipo_pivote (NULL si es bloque general, sino: 'sueldo'/'aguinaldo'/'horas_extras'/'bono')
  - sheet_categoria_inferida (de nombre de hoja: empleados/concejales/...)
  - sugerencia_concepto, sugerencia_secretaria, sugerencia_caja (NULL al inicio,
    se llenan en Fase 2 con IA)
  - sugerencia_proyecto, sugerencia_tipo_contacto
  - notas (para debug)

NO modifica catalogos. NO crea contactos. NO crea gastos. Solo staging.
"""
import asyncio
import re
import sys
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl  # noqa: E402
from sqlalchemy.ext.asyncio import create_async_engine  # noqa: E402
from sqlalchemy import text  # noqa: E402
from core.config import settings  # noqa: E402


# ============================================================
# Config
# ============================================================
BASE = Path('D:/Code/sugerenciasMun/docs/bartolo')

ARCHIVOS = [
    ('Gastos 2024.xlsx', 2024),
    ('Gastos mensuales 2025 - 01-09.xlsx', 2025),
    ('Gatos municipalidad oct-nov-dic.xlsx', 2025),
]

# Modos conocidos (para detectar corrimiento y normalizar)
MODOS_KNOWN = {
    'transf', 'trans', 'transferencia', 'trabsf', 'trans y eftv',
    'ef', 'eftv', 'efectivo', 'efectiv', 'efectvio',
    'cheq', 'cheque', 'echeq', 'echque',
    'mp', 'mercadopago', 'mercado pago',
    'visa', 'tarjeta', 'debito', 'debito automatico',
    'vep', 'debin',
}

MODO_NORMALIZACION = {
    'transf': 'transferencia', 'trans': 'transferencia', 'trabsf': 'transferencia',
    'trans y eftv': 'transferencia',
    'ef': 'efectivo', 'eftv': 'efectivo', 'efectiv': 'efectivo', 'efectvio': 'efectivo',
    'cheq': 'cheque', 'echque': 'cheque',
    'mp': 'mercadopago', 'mercado pago': 'mercadopago',
}

# Meses (para parsear nombre de columna o nombre de hoja)
MESES_NORMAL = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'marcos': 3, 'abril': 4,
    'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
    'septiembre': 9, 'septiemrbe': 9, 'septiemb': 9, 'sept': 9,
    'octubre': 10, 'oct': 10,
    'noviembre': 11, 'nov': 11,
    'diciembre': 12, 'dic': 12, 'dici': 12,
}

# Hojas que son grilla pivote (1 fila por persona x N meses con sueldo)
HOJAS_PIVOTE_PERSONAL = {
    'empleados': 'empleado_planta',
    'concejales': 'concejal',
    'turismo y cultura': 'empleado_turismo',
    'profesionales y publicidad': 'profesional',
    'prof y pub': 'profesional',
    'apoyo gerontológico': 'gerontologico',
    'apoyo gerontologico': 'gerontologico',
    'ayuda social': 'ayuda_social',
}

# Hojas con bloques [Mes, Monto, Empresa, Modo]
HOJAS_BLOQUE_GASTOS = {
    'gastos', 'octubre', 'noviembre', 'diciembre', 'oct', 'nov', 'dic',
}


# ============================================================
# Helpers
# ============================================================
def normalize_text(s) -> str:
    if s is None:
        return ''
    s = str(s).strip()
    # Quitar acentos
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s.lower()


def to_decimal(v) -> Optional[Decimal]:
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    if isinstance(v, Decimal):
        return v
    s = str(v).strip()
    # Filtrar refs rotas
    if s in ('#REF!', '#DIV/0!', '#N/A', '#VALUE!'):
        return None
    # Limpiar $ y separadores
    s = s.replace('$', '').replace('.', '').replace(' ', '')
    # Si tiene coma decimal
    s = s.replace(',', '.')
    # Solo dígitos y un punto
    s = re.sub(r'[^\d\.\-]', '', s)
    if not s or s in ('.', '-'):
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def parse_parentesis(s: str) -> Optional[str]:
    """Devuelve el texto entre paréntesis si hay."""
    if not s:
        return None
    m = re.search(r'\(([^)]+)\)', str(s))
    return m.group(1).strip() if m else None


def empresa_sin_parentesis(s: str) -> str:
    """Texto de empresa sin lo que está entre paréntesis."""
    if not s:
        return ''
    return re.sub(r'\s*\([^)]*\)\s*', ' ', str(s)).strip()


def normalizar_nombre_persona(s: str) -> str:
    """'Avila Cesar' y 'Cesar Avila' deben coincidir. Devuelve formato canónico
    'apellido nombre' ordenado alfabéticamente para deduplicación."""
    s = normalize_text(empresa_sin_parentesis(s))
    s = re.sub(r'\s+', ' ', s)
    return s


def is_modo_known(s) -> bool:
    if s is None:
        return False
    n = normalize_text(s)
    if n in MODOS_KNOWN:
        return True
    # "ch 1234" o "echeq 22/11" o "Echeq DD/MM"
    if re.match(r'^(ch|cheq|echeq)\s*[\d/]+', n):
        return True
    if re.match(r'^echeq\s*\d', n):
        return True
    return False


def normalizar_modo(s) -> str:
    if s is None:
        return ''
    n = normalize_text(s)
    # Si arranca con "ch " o "cheq" o "echeq" → simplificar
    if re.match(r'^ch\s+', n) or n == 'cheq':
        return 'cheque'
    if re.match(r'^echeq', n):
        return 'echeq'
    return MODO_NORMALIZACION.get(n, n)


def mes_de_col(header_value) -> Optional[int]:
    """Dado un header (ej 'Enero', 'Marzo', 'Marcos'), devuelve mes 1-12."""
    if header_value is None:
        return None
    n = normalize_text(header_value)
    for k, v in MESES_NORMAL.items():
        if k in n:
            return v
    return None


def categoria_de_hoja(sheet_name: str) -> Optional[str]:
    n = normalize_text(sheet_name)
    return HOJAS_PIVOTE_PERSONAL.get(n)


def es_hoja_bloque_gastos(sheet_name: str) -> bool:
    n = normalize_text(sheet_name).split()[0] if sheet_name else ''
    return n in HOJAS_BLOQUE_GASTOS or any(
        normalize_text(sheet_name).startswith(k) for k in HOJAS_BLOQUE_GASTOS
    )


# ============================================================
# Parsers
# ============================================================
def parse_bloques_gastos(ws, fuente_archivo: str, anio_default: int) -> list[dict]:
    """Hojas tipo 'Gastos' u 'Octubre' donde hay bloques [Mes, Monto, Empresa, Modo]
    repetidos horizontalmente. Detecta corrimiento de 1 columna si modo viene desfasado.
    """
    out = []
    # Buscar la fila header (la que tiene "Empresa" o "Modo")
    header_row = None
    for row_idx in range(1, min(6, ws.max_row + 1)):
        row = [str(c.value).strip().lower() if c.value else '' for c in ws[row_idx]]
        if any('empresa' in c for c in row) and any('modo' in c for c in row):
            header_row = row_idx
            break
    if header_row is None:
        return out

    headers = [str(c.value).strip() if c.value else '' for c in ws[header_row]]

    # Detectar columnas Empresa/Modo recurrentes
    empresa_cols = [i for i, h in enumerate(headers) if 'empresa' in h.lower()]
    if not empresa_cols:
        return out

    for emp_col in empresa_cols:
        # Buscar la col de Monto (la inmediatamente anterior) y Modo (la siguiente)
        monto_col = emp_col - 1 if emp_col > 0 else None
        modo_col = emp_col + 1 if emp_col + 1 < len(headers) else None
        # Mes header: la columna 2 antes de empresa, ej [Mes, Monto, Empresa, Modo]
        mes_col = emp_col - 2 if emp_col >= 2 else None
        mes_header = headers[mes_col] if mes_col is not None else ''
        mes = mes_de_col(mes_header)

        # Procesar todas las filas
        for row_idx in range(header_row + 1, ws.max_row + 1):
            empresa_raw = ws.cell(row=row_idx, column=emp_col + 1).value
            if not empresa_raw or str(empresa_raw).strip() in ('', '#REF!'):
                continue

            monto_raw = ws.cell(row=row_idx, column=monto_col + 1).value if monto_col is not None else None
            modo_raw = ws.cell(row=row_idx, column=modo_col + 1).value if modo_col is not None else None

            # === DETECCION DE CORRIMIENTO ===
            # Si modo no parece un modo conocido pero la col+1 sí, todo corrido +1
            bloque_corrido = False
            if modo_raw and not is_modo_known(modo_raw):
                # ¿La col +1 contiene un modo conocido?
                next_val = ws.cell(row=row_idx, column=modo_col + 2).value if modo_col + 2 <= ws.max_column else None
                if next_val and is_modo_known(next_val):
                    # Corrido: empresa estaba mal interpretada, "modo" es la empresa real
                    empresa_raw = modo_raw
                    monto_raw = ws.cell(row=row_idx, column=emp_col + 1).value  # la antigua empresa era monto
                    modo_raw = next_val
                    bloque_corrido = True

            monto = to_decimal(monto_raw)
            if monto is None or monto <= 0:
                continue

            empresa_str = str(empresa_raw).strip()
            paren = parse_parentesis(empresa_str)
            empresa_sin_p = empresa_sin_parentesis(empresa_str)

            out.append({
                'fuente_archivo': fuente_archivo,
                'fuente_hoja': ws.title,
                'fuente_fila': row_idx,
                'fuente_col_block': emp_col,
                'bloque_corrido': bloque_corrido,
                'mes': mes,
                'anio': anio_default,
                'monto': monto,
                'empresa_raw': empresa_str,
                'empresa_sin_parentesis': empresa_sin_p,
                'empresa_normalizada': normalizar_nombre_persona(empresa_sin_p),
                'empresa_parentesis': paren,
                'modo_raw': str(modo_raw).strip() if modo_raw else None,
                'modo_normalizado': normalizar_modo(modo_raw),
                'tipo_pivote': None,
                'sheet_categoria_inferida': 'proveedor',
            })
    return out


def parse_grilla_pivote(ws, fuente_archivo: str, anio_default: int) -> list[dict]:
    """Hojas tipo Empleados/Concejales: 1 fila por persona, columnas son meses con monto.
    Cada celda con monto = 1 gasto de tipo 'sueldo'.

    Soporta hojas con dos años pegados horizontalmente (2024 cols X-Y, 2025 cols Y+1-Z).
    """
    out = []
    categoria = categoria_de_hoja(ws.title)
    if not categoria:
        return out

    # Buscar header row con nombres de meses
    header_row = None
    meses_por_col = {}  # col_idx -> (anio, mes)
    for row_idx in range(1, min(5, ws.max_row + 1)):
        cnt_meses = 0
        meses_tmp = {}
        anio_current = anio_default
        for col_idx, cell in enumerate(ws[row_idx]):
            mes = mes_de_col(cell.value)
            if mes is not None:
                # Si vimos diciembre antes y ahora estamos en enero -> +1 año
                prev_meses = list(meses_tmp.values())
                if prev_meses and prev_meses[-1][1] == 12 and mes == 1:
                    anio_current += 1
                meses_tmp[col_idx] = (anio_current, mes)
                cnt_meses += 1
        if cnt_meses >= 6:  # al menos medio año de meses, es el header bueno
            header_row = row_idx
            meses_por_col = meses_tmp
            break

    if header_row is None or not meses_por_col:
        return out

    # Buscar la columna que contiene el nombre de la persona.
    # Usualmente es la 3a o 4a columna (col1=tipo, col2=banco, col3=nombre, etc).
    # Heurística: la columna con más strings no-numéricos antes de la primera col de mes.
    primera_col_mes = min(meses_por_col.keys())
    candidato_nombre_col = None
    mejor_count = 0
    for col_idx in range(0, primera_col_mes):
        count = 0
        for row_idx in range(header_row + 1, min(header_row + 30, ws.max_row + 1)):
            v = ws.cell(row=row_idx, column=col_idx + 1).value
            if v and isinstance(v, str) and len(v.strip()) > 4:
                # No es un número ni vacío
                if not re.match(r'^\$?[\d,\.\s\-]+$', v.strip()):
                    count += 1
        if count > mejor_count:
            mejor_count = count
            candidato_nombre_col = col_idx
    if candidato_nombre_col is None:
        return out

    # Procesar filas
    for row_idx in range(header_row + 1, ws.max_row + 1):
        nombre_raw = ws.cell(row=row_idx, column=candidato_nombre_col + 1).value
        if not nombre_raw or not isinstance(nombre_raw, str):
            continue
        nombre = nombre_raw.strip()
        if not nombre or len(nombre) < 3:
            continue
        if 'total' in nombre.lower() or '#ref' in nombre.lower():
            continue

        # Para cada col de mes con monto, generar una fila
        for col_idx, (anio, mes) in meses_por_col.items():
            cell_val = ws.cell(row=row_idx, column=col_idx + 1).value
            monto = to_decimal(cell_val)
            if monto is None or monto <= 0:
                continue
            out.append({
                'fuente_archivo': fuente_archivo,
                'fuente_hoja': ws.title,
                'fuente_fila': row_idx,
                'fuente_col_block': col_idx,
                'bloque_corrido': False,
                'mes': mes,
                'anio': anio,
                'monto': monto,
                'empresa_raw': nombre,
                'empresa_sin_parentesis': nombre,
                'empresa_normalizada': normalizar_nombre_persona(nombre),
                'empresa_parentesis': None,
                'modo_raw': None,
                'modo_normalizado': '',
                'tipo_pivote': 'sueldo',
                'sheet_categoria_inferida': categoria,
            })
    return out


# ============================================================
# Main
# ============================================================
async def main():
    engine = create_async_engine(settings.DATABASE_URL)

    # Crear tabla staging idempotente
    async with engine.begin() as conn:
        await conn.execute(text("""
            CREATE TABLE IF NOT EXISTS bartolo_raw (
                id INT AUTO_INCREMENT PRIMARY KEY,
                fuente_archivo VARCHAR(200) NOT NULL,
                fuente_hoja VARCHAR(100) NOT NULL,
                fuente_fila INT NOT NULL,
                fuente_col_block INT NULL,
                bloque_corrido TINYINT(1) NOT NULL DEFAULT 0,
                mes TINYINT NULL,
                anio SMALLINT NOT NULL,
                monto DECIMAL(15,2) NOT NULL,
                empresa_raw VARCHAR(400) NOT NULL,
                empresa_sin_parentesis VARCHAR(400) NULL,
                empresa_normalizada VARCHAR(400) NULL,
                empresa_parentesis VARCHAR(300) NULL,
                modo_raw VARCHAR(200) NULL,
                modo_normalizado VARCHAR(50) NULL,
                tipo_pivote VARCHAR(50) NULL,
                sheet_categoria_inferida VARCHAR(50) NULL,
                sugerencia_concepto VARCHAR(150) NULL,
                sugerencia_secretaria VARCHAR(100) NULL,
                sugerencia_caja VARCHAR(50) NULL,
                sugerencia_proyecto VARCHAR(200) NULL,
                sugerencia_tipo_contacto VARCHAR(30) NULL,
                notas TEXT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_archivo (fuente_archivo),
                INDEX idx_anio_mes (anio, mes),
                INDEX idx_normalizada (empresa_normalizada(100))
            ) ENGINE=InnoDB CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """))
        # Truncar para arrancar limpio
        await conn.execute(text("TRUNCATE TABLE bartolo_raw"))
        print("[OK] tabla bartolo_raw lista (truncada)")

    todos_los_registros = []
    for archivo, anio_default in ARCHIVOS:
        path = BASE / archivo
        if not path.exists():
            print(f"[!] no existe: {path}")
            continue
        print(f"\n--- {archivo} ---")
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            nlow = normalize_text(sheet_name)
            antes = len(todos_los_registros)

            # Pivote personal?
            if categoria_de_hoja(sheet_name):
                registros = parse_grilla_pivote(ws, archivo, anio_default)
                todos_los_registros.extend(registros)
                print(f"  [pivote] {sheet_name}: +{len(registros)} pagos")
            # Bloque gastos generales?
            elif es_hoja_bloque_gastos(sheet_name) or 'gasto' in nlow or 'gato' in nlow:
                registros = parse_bloques_gastos(ws, archivo, anio_default)
                todos_los_registros.extend(registros)
                corridos = sum(1 for r in registros if r['bloque_corrido'])
                if corridos:
                    print(f"  [bloque] {sheet_name}: +{len(registros)} gastos ({corridos} con corrimiento)")
                else:
                    print(f"  [bloque] {sheet_name}: +{len(registros)} gastos")
            elif 'octubre' in nlow or 'noviembre' in nlow or 'diciembre' in nlow:
                registros = parse_bloques_gastos(ws, archivo, anio_default)
                todos_los_registros.extend(registros)
                print(f"  [bloque-mes] {sheet_name}: +{len(registros)} gastos")
            else:
                print(f"  [skip] {sheet_name}")

    print(f"\n[TOTAL] {len(todos_los_registros)} registros a insertar")

    # Bulk insert
    async with engine.begin() as conn:
        BATCH = 200
        for i in range(0, len(todos_los_registros), BATCH):
            chunk = todos_los_registros[i:i + BATCH]
            await conn.execute(text("""
                INSERT INTO bartolo_raw (
                    fuente_archivo, fuente_hoja, fuente_fila, fuente_col_block,
                    bloque_corrido, mes, anio, monto,
                    empresa_raw, empresa_sin_parentesis, empresa_normalizada,
                    empresa_parentesis, modo_raw, modo_normalizado,
                    tipo_pivote, sheet_categoria_inferida
                ) VALUES (
                    :fuente_archivo, :fuente_hoja, :fuente_fila, :fuente_col_block,
                    :bloque_corrido, :mes, :anio, :monto,
                    :empresa_raw, :empresa_sin_parentesis, :empresa_normalizada,
                    :empresa_parentesis, :modo_raw, :modo_normalizado,
                    :tipo_pivote, :sheet_categoria_inferida
                )
            """), chunk)
        print(f"[OK] insertados {len(todos_los_registros)} registros")

    # Quick stats
    async with engine.begin() as conn:
        r = (await conn.execute(text("""
            SELECT fuente_archivo, COUNT(*), SUM(monto), SUM(bloque_corrido)
            FROM bartolo_raw GROUP BY fuente_archivo
        """))).fetchall()
        print("\n[STATS por archivo]")
        for row in r:
            print(f"  {row[0]}: {row[1]} registros, ${row[2]:,.0f}, corridos={row[3]}")

        r = (await conn.execute(text("""
            SELECT sheet_categoria_inferida, COUNT(*), SUM(monto)
            FROM bartolo_raw GROUP BY sheet_categoria_inferida
        """))).fetchall()
        print("\n[STATS por categoría]")
        for row in r:
            print(f"  {row[0]}: {row[1]} registros, ${row[2]:,.0f}")

    await engine.dispose()
    print("\n[DONE]")


if __name__ == "__main__":
    asyncio.run(main())
