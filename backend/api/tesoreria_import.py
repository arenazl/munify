"""Importadores para Tesoreria.

- POST /tesoreria/import/excel-matriz   importa el Excel del intendente
                                         (formato matriz: filas=personas,
                                         columnas=meses).
- POST /tesoreria/import/kmz             actualiza lat/lon de contactos
                                         existentes haciendo match por nombre.

Ambos endpoints requieren admin.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Optional
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from core.database import get_db
from core.security import get_current_user
from core.tenancy import get_effective_municipio_id
from models import (
    Contacto, Gasto, GastoCuota, User, RolUsuario,
    TipoContacto, DestinoGasto, TipoFinanciacion, FormaPago, EstadoGastoCuota,
)

router = APIRouter()


# Mapeo de sheet del Excel del intendente -> tipo de contacto y subtipo
SHEET_TO_TIPO = {
    "concejales": (TipoContacto.CONCEJAL, None),
    "empleados": (TipoContacto.EMPLEADO, None),
    "profesionales y publicidad": (TipoContacto.PROFESIONAL, None),
    "turismo y cultura": (TipoContacto.BENEFICIARIO, "turismo y cultura"),
    "obra vestuarios": (TipoContacto.PROVEEDOR, "obra"),
}

MES_A_NUMERO = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def _require_admin(user: User):
    if user.rol != RolUsuario.ADMIN:
        raise HTTPException(status_code=403, detail="Solo admin puede importar")


def _normalize(s: str) -> str:
    """Lowercase + strip + sin tildes basicas — para matching de nombres."""
    s = (s or "").strip().lower()
    repl = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "ñ": "n", "Á": "a", "É": "e", "Í": "i", "Ó": "o", "Ú": "u",
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    return s


# ============================================================
# Excel matriz importer
# ============================================================

@router.post("/excel-matriz")
async def importar_excel_matriz(
    request: Request,
    archivo: UploadFile = File(..., description="xlsx con formato del intendente"),
    anio: int = 2026,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Importa el Excel del intendente.

    Por cada sheet conocido (Concejales / Empleados / Profesionales y
    publicidad / Turismo y cultura / Obra Vestuarios):
      - Detecta la columna `Alias` (si existe) y las columnas de meses
        (Enero, Febrero, ..., Diciembre, Aguinaldo, Medio aguinaldo).
      - Por cada fila con nombre: crea/actualiza Contacto.
      - Por cada celda con monto > 0: crea un Gasto (CONTADO) con concepto
        = nombre de la columna y fecha = primer dia del mes correspondiente.

    Devuelve resumen { contactos_creados, gastos_creados, sheets_procesadas }.
    """
    _require_admin(current_user)
    municipio_id = get_effective_municipio_id(request, current_user)
    if not municipio_id:
        raise HTTPException(status_code=400, detail="Municipio no resuelto")

    contenido = await archivo.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel invalido: {e}")

    contactos_creados = 0
    contactos_actualizados = 0
    gastos_creados = 0
    sheets_procesadas: List[str] = []
    sheets_ignoradas: List[str] = []

    # Pre-cargar contactos existentes del municipio para matching por nombre
    res = await db.execute(
        select(Contacto).where(Contacto.municipio_id == municipio_id)
    )
    existentes: Dict[str, Contacto] = {
        _normalize(c.nombre + " " + (c.apellido or "")): c
        for c in res.scalars().all()
    }

    for sheet_name in wb.sheetnames:
        key = _normalize(sheet_name)
        if key not in SHEET_TO_TIPO:
            sheets_ignoradas.append(sheet_name)
            continue
        tipo, subtipo = SHEET_TO_TIPO[key]
        ws = wb[sheet_name]

        # Leer header
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            sheets_ignoradas.append(sheet_name)
            continue

        # Mapear columnas: nombre (col 0 o donde haya texto), alias, meses
        # En el Excel real el alias suele estar en col 1 (Concejales) o col 3 (Empleados).
        col_alias = None
        col_concepto: Dict[int, str] = {}   # idx_col -> "Enero" / "Aguinaldo" / etc.
        col_nombre = 0   # default

        for idx, h in enumerate(header):
            if h is None:
                continue
            h_norm = _normalize(str(h))
            if "alias" in h_norm:
                col_alias = idx
            elif h_norm in MES_A_NUMERO:
                col_concepto[idx] = f"Sueldo {h_norm.capitalize()}"
            elif h_norm == "aguinaldo":
                col_concepto[idx] = "Aguinaldo"
            elif "medio aguinaldo" in h_norm:
                col_concepto[idx] = "Medio aguinaldo"
            elif "horas extras" in h_norm:
                col_concepto[idx] = "Horas extras"

        # Encontrar la columna del nombre: primera celda no-vacia en row 0
        # que no sea None y no sea un mes/alias.
        for idx, h in enumerate(header):
            if h is None:
                continue
            h_norm = _normalize(str(h))
            if h_norm in MES_A_NUMERO:
                break
            if "alias" in h_norm:
                continue
            if h_norm in ("", "monto", "modo", "empresa"):
                continue
            # Probable columna de nombre — pero solo si no hemos confirmado otra
            # Heuristic: en muchos sheets la primera col del header es None y los
            # nombres aparecen en col 0. En Empleados va en col 2.
        # Simpler heuristic: probar col 0; si la mayoria de filas tienen str ahi, usar 0.
        # Si no, probar col 2.
        sample_rows = list(ws.iter_rows(min_row=2, max_row=10, values_only=True))
        def es_nombre(v):
            return isinstance(v, str) and len(v.strip()) > 2 and not v.strip().isdigit()
        scores = [0, 0, 0, 0]
        for r in sample_rows:
            for i in range(min(4, len(r))):
                if es_nombre(r[i]):
                    scores[i] += 1
        col_nombre = scores.index(max(scores)) if max(scores) > 0 else 0

        # Procesar filas
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or col_nombre >= len(row):
                continue
            nombre_raw = row[col_nombre]
            if not nombre_raw or not isinstance(nombre_raw, str):
                continue
            nombre_raw = nombre_raw.strip()
            if not nombre_raw or nombre_raw.lower().startswith(("total", "subtotal", "gastos:")):
                continue

            # Split nombre / apellido (heuristic: 2 palabras = nombre apellido)
            partes = nombre_raw.split()
            if len(partes) >= 2:
                nombre = partes[0]
                apellido = " ".join(partes[1:])
            else:
                nombre = nombre_raw
                apellido = None

            alias = None
            if col_alias is not None and col_alias < len(row):
                v = row[col_alias]
                if v and str(v).strip():
                    alias = str(v).strip()[:60]

            # Crear o actualizar contacto
            key_match = _normalize(nombre_raw)
            contacto = existentes.get(key_match)
            if contacto is None:
                contacto = Contacto(
                    municipio_id=municipio_id,
                    nombre=nombre,
                    apellido=apellido,
                    alias_pago=alias,
                    tipo=tipo,
                    subtipo=subtipo,
                )
                db.add(contacto)
                await db.flush()
                existentes[key_match] = contacto
                contactos_creados += 1
            else:
                # Actualizar alias si no lo tenia
                if alias and not contacto.alias_pago:
                    contacto.alias_pago = alias
                if not contacto.tipo:
                    contacto.tipo = tipo
                contactos_actualizados += 1

            # Crear gastos por cada celda de mes con monto > 0
            for col_idx, concepto in col_concepto.items():
                if col_idx >= len(row):
                    continue
                monto = row[col_idx]
                if not isinstance(monto, (int, float)) or monto <= 0:
                    continue

                # Mes -> dia 1
                mes = MES_A_NUMERO.get(_normalize(concepto.replace("Sueldo ", "").strip()))
                if not mes:
                    if "aguinaldo" in concepto.lower():
                        mes = 12 if "medio" not in concepto.lower() else 6
                    elif "horas extras" in concepto.lower():
                        mes = 1
                    else:
                        mes = 1
                fecha_gasto = date(anio, mes, 1)

                gasto = Gasto(
                    municipio_id=municipio_id,
                    creador_id=current_user.id,
                    destino_tipo=DestinoGasto.CONTACTO,
                    destino_contacto_id=contacto.id,
                    concepto=concepto,
                    monto_pesos=Decimal(str(monto)),
                    fecha=fecha_gasto,
                    tipo_financiacion=TipoFinanciacion.CONTADO,
                    forma_pago=FormaPago.TRANSFERENCIA,
                )
                db.add(gasto)
                await db.flush()
                # 1 cuota pagada
                db.add(GastoCuota(
                    gasto_id=gasto.id,
                    numero=1,
                    monto=gasto.monto_pesos,
                    fecha_vencimiento=fecha_gasto,
                    fecha_pago=fecha_gasto,
                    estado=EstadoGastoCuota.PAGADA,
                    forma_pago=FormaPago.TRANSFERENCIA,
                ))
                gastos_creados += 1

        sheets_procesadas.append(sheet_name)

    await db.commit()
    return {
        "ok": True,
        "contactos_creados": contactos_creados,
        "contactos_actualizados": contactos_actualizados,
        "gastos_creados": gastos_creados,
        "sheets_procesadas": sheets_procesadas,
        "sheets_ignoradas": sheets_ignoradas,
    }


# ============================================================
# KMZ importer
# ============================================================

KML_NS = "{http://www.opengis.net/kml/2.2}"


@router.post("/kmz")
async def importar_kmz(
    request: Request,
    archivo: UploadFile = File(..., description="KMZ con placemarks geolocalizados"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Importa coordenadas desde un KMZ.

    Por cada <Placemark> del KML extrae <name> y <coordinates>. Busca un
    Contacto del municipio con nombre que matchee y le actualiza lat/lon.

    NO crea contactos nuevos (solo actualiza). Si querias crear, primero
    corre el import-excel-matriz.
    """
    _require_admin(current_user)
    municipio_id = get_effective_municipio_id(request, current_user)
    if not municipio_id:
        raise HTTPException(status_code=400, detail="Municipio no resuelto")

    contenido = await archivo.read()
    try:
        with ZipFile(BytesIO(contenido)) as zf:
            kml_names = [n for n in zf.namelist() if n.lower().endswith(".kml")]
            if not kml_names:
                raise HTTPException(status_code=400, detail="KMZ sin archivo .kml adentro")
            kml_data = zf.read(kml_names[0])
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"KMZ invalido: {e}")

    try:
        root = ET.fromstring(kml_data)
    except ET.ParseError as e:
        raise HTTPException(status_code=400, detail=f"KML invalido: {e}")

    # Cargar contactos del muni para matching
    res = await db.execute(
        select(Contacto).where(Contacto.municipio_id == municipio_id)
    )
    by_name: Dict[str, Contacto] = {
        _normalize(c.nombre + " " + (c.apellido or "")): c
        for c in res.scalars().all()
    }

    actualizados = 0
    no_matcheados: List[str] = []

    for placemark in root.iter(f"{KML_NS}Placemark"):
        name_el = placemark.find(f"{KML_NS}name")
        coords_el = None
        # Point/coordinates o cualquier coordinates
        for c in placemark.iter(f"{KML_NS}coordinates"):
            coords_el = c
            break
        if name_el is None or coords_el is None or not coords_el.text:
            continue
        nombre = (name_el.text or "").strip()
        coords_text = coords_el.text.strip().split()[0]  # "lon,lat,alt"
        try:
            partes = coords_text.split(",")
            lon = float(partes[0])
            lat = float(partes[1])
        except (ValueError, IndexError):
            continue

        key = _normalize(nombre)
        contacto = by_name.get(key)
        if contacto:
            contacto.latitud = lat
            contacto.longitud = lon
            actualizados += 1
        else:
            no_matcheados.append(nombre)

    await db.commit()
    return {
        "ok": True,
        "actualizados": actualizados,
        "no_matcheados": no_matcheados,
    }
